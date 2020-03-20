"""This python project uses twilio services to send invitations to a list of phone numbers"""

from twilio.rest import Client
from twilio.base.exceptions import TwilioRestException
import openpyxl
import re
import WeddingData             # Data file includes twilio account, location of excel, invite body and the sender's name

"""Global Variables"""
account_sid = WeddingData.sid
auth_token = WeddingData.token
client = Client(account_sid, auth_token)
loc = WeddingData.loc
invites = openpyxl.load_workbook(loc)
invites_sheet = invites['Sheet1']


def good_number(phone_number):
    """
    :param phone_number:
    :return: True if phone number is in correct format
    """
    pattern = re.compile(r'\+972[\d]{9}')
    if pattern.fullmatch(phone_number):
        return True
    else:
        return False


def format_number(phone_number):
    """
    :param phone_number:
    :return: The correct format to be used to send the message
    """
    leading_zero = re.compile(r'0\d{9}')
    three_dash = re.compile(r'\d{3}-\d{7}')
    two_dash = re.compile(r'[^0]\d-\d{7}')
    if leading_zero.fullmatch(phone_number):
        return '+972' + phone_number[1:]
    elif three_dash.fullmatch(phone_number):
        return '+972' + phone_number[1:3] + phone_number[4:]
    elif two_dash.fullmatch(phone_number):
        return '+972' + phone_number[0:2] + phone_number[3:]
    else:
        return None


def send_message(my_client, phone_number):
    """
    :param my_client:
    :param phone_number:
    :return: sid of message or message error
    """
    try:
        message = my_client.messages.create(body=WeddingData.body,
                                            from_=WeddingData.from_,
                                            to=phone_number)
        return True, message
    except TwilioRestException as e:
        return False, e


def get_max_row():
    """
    :return: number of phone numbers in excel
    """
    counter = 0
    i = 1
    while True:
        cell = invites_sheet.cell(row=i, column=2).value
        if cell is not None:
            counter += 1
            i += 1
        else:
            return counter


def main():
    max_row = get_max_row()
    for i in range(2, max_row + 1):
        number = invites_sheet.cell(row=i, column=2).value  # Get phone number
        sent_cell = invites_sheet.cell(row=i, column=3)
        sid_cell = invites_sheet.cell(row=i, column=4)
        error_cell = invites_sheet.cell(row=i, column=5)
        if not good_number(number):                         # Check if number is in +972 format
            number = format_number(number)
        if number is None:                                  # If not in format, write error and skip
            sent_cell.value = 'No'
            error_cell.value = 'Phone number is wrong'
            invites.save(loc)
            continue
        (sent, message) = send_message(client, number)      # Send message
        if sent is True:
            sent_cell.value = 'Yes'
            sid_cell.value = message.sid
        else:
            sent_cell.value = 'No'
            error_cell.value = message
        invites.save(loc)


if __name__ == '__main__':
    main()
